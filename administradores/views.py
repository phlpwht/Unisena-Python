from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.cache import never_cache
from usuario.models import Usuario, Rol
from usuario.utils import _get_user_block_status
from uniformes.models import Pedido, EstadoPedido, Prendas, CalificacionPedido, DetallePedido
from locales.models import Local, Notificacion
from django.db.models import Q, Sum, Count, DateField, OuterRef, Subquery
from django.db.models.functions import TruncMonth, TruncDay, TruncYear
from django.db import transaction  # Importar transaction
from django.contrib import messages
from django.urls import reverse
from django.middleware.csrf import get_token
from django.utils.timezone import now, timedelta
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from decimal import Decimal
from datetime import datetime, time
import io

@never_cache
def dashadmin(request):
    # Verificación de seguridad: Sesión y Rol de Administrador
    if "usuario_id" not in request.session:
        return redirect("login")

    # Verificamos que sea Administrador
    if request.session.get("usuario_rol") != "Administrador":
        return redirect("login")

    # Verificar si el administrador está bloqueado (reutiliza la lógica central)
    usuario_bloqueado, msg_bloqueo, user_obj = _get_user_block_status(request)
    if usuario_bloqueado:
        # Reutilizamos la plantilla `politicas.html` que ya muestra el aviso de cuenta restringida
        return render(request, "politicas.html", {
            "usuario_bloqueado": usuario_bloqueado,
            "msg_bloqueo": msg_bloqueo,
            "nombre": request.session.get("usuario_nombre"),
            "rol": request.session.get("usuario_rol"),
        })

    # Detectar qué sección mostrar (por defecto 'inicio')
    section = request.GET.get('section', 'dashboard') or 'dashboard'
    print("DEBUG request.GET:", dict(request.GET))

    # Procesar acciones POST (Eliminar o Cambiar Rol)
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        accion = request.POST.get("accion")

        if accion == "eliminar":
            with transaction.atomic(): # Usamos una transacción para asegurar la consistencia
                usuario_eliminar = get_object_or_404(Usuario, id=user_id)
                
                # Evitar que el admin se elimine a sí mismo por error
                if str(usuario_eliminar.id) == str(request.session.get("usuario_id")):
                    messages.error(request, "No puedes eliminar tu propia cuenta administrativa.")
                    return redirect(f"{reverse('admin_dashboard')}?section=usuarios")
                
                # Validar si el usuario tiene pedidos pendientes o en proceso (Cliente)
                if usuario_eliminar.rol.nombre_rol == "Cliente":
                    tiene_pedidos_cliente = Pedido.objects.filter(
                        usuario=usuario_eliminar,
                        estado__estado_pedido__in=['PENDIENTE', 'En Proceso']
                    ).exists()
                    if tiene_pedidos_cliente:
                        messages.error(request, f"No se puede eliminar a {usuario_eliminar.nombres} porque tiene pedidos pendientes o en proceso como cliente.")
                        return redirect(f"{reverse('admin_dashboard')}?section=usuarios")
                
                # Validar si el usuario tiene pedidos pendientes o en proceso (Vendedor)
                if usuario_eliminar.rol.nombre_rol == "Vendedor":
                    tiene_pedidos_vendedor = Pedido.objects.filter(
                        detallepedido__prenda__idLocal__IdUsuario=usuario_eliminar,
                        estado__estado_pedido__in=['PENDIENTE', 'En Proceso']
                    ).exists()
                    if tiene_pedidos_vendedor:
                        messages.error(request, f"No se puede eliminar a {usuario_eliminar.nombres} porque tiene pedidos pendientes o en proceso asociados a sus locales.")
                        return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

                # Si todas las validaciones pasan, procedemos a eliminar y guardar para deshacer
                request.session['last_deleted_user'] = {
                    'nombres': usuario_eliminar.nombres,
                    'apellidos': usuario_eliminar.apellidos,
                    'correo': usuario_eliminar.correo,
                    'fecha_nacimiento': str(usuario_eliminar.fecha_nacimiento),
                    'tipo_identificacion': usuario_eliminar.tipo_identificacion,
                    'num_identificacion': usuario_eliminar.num_identificacion,
                    'password': usuario_eliminar.password,
                    'rol_id': usuario_eliminar.rol_id
                }
                usuario_eliminar.delete()
                messages.success(request, f"Usuario <b>{usuario_eliminar.nombres}</b> eliminado correctamente. ¿Deseas deshacer esta acción? <form method='POST' class='inline ml-2'><input type='hidden' name='csrfmiddlewaretoken' value='{get_token(request)}'><input type='hidden' name='accion' value='deshacer'><button type='submit' class='underline font-black uppercase text-emerald-700 hover:text-emerald-900'>[Deshacer]</button></form> <span class='countdown-timer ml-2 opacity-50'>(10s)</span>")
            return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

        elif accion == "deshacer":
            user_data = request.session.get('last_deleted_user')
            if user_data:
                Usuario.objects.create(**user_data)
                del request.session['last_deleted_user']
                messages.success(request, "♻️ Usuario restaurado con éxito.")
                return redirect(f"{reverse('admin_dashboard')}?section=usuarios")
            
            local_data = request.session.get('last_deleted_local')
            if local_data:
                Local.objects.create(**local_data)
                del request.session['last_deleted_local']
                messages.success(request, "♻️ Local restaurado con éxito.")
                return redirect(f"{reverse('admin_dashboard')}?section=locales")

            prenda_data = request.session.get('last_deleted_prenda')
            if prenda_data:
                Prendas.objects.create(**prenda_data)
                del request.session['last_deleted_prenda']
                messages.success(request, "♻️ Uniforme restaurado con éxito.")
                return redirect(f"{reverse('admin_dashboard')}?section=uniformes")
            return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

        elif accion == "desbloquear_usuario":
            usuario_id = request.POST.get("user_id")
            usuario = get_object_or_404(Usuario, id=usuario_id)
            usuario.bloqueado_hasta = None
            usuario.pendiente_aceptar_politicas = False
            usuario.save()
            messages.success(request, f"Usuario {usuario.nombres} desbloqueado correctamente.")
            return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

        elif accion == "cambiar_rol":
            nuevo_rol_id = request.POST.get("rol_id")
            usuario_editar = get_object_or_404(Usuario, id=user_id)
            nuevo_rol = get_object_or_404(Rol, id=nuevo_rol_id)
            
            # --- VALIDACIONES DE NEGOCIO ---
            # 1. Si es Vendedor y tiene pedidos en sus locales
            if usuario_editar.rol.nombre_rol == "Vendedor":
                tiene_pedidos_vendedor = Pedido.objects.filter(
                    detallepedido__prenda__idLocal__IdUsuario=usuario_editar,
                    estado__estado_pedido__in=['PENDIENTE', 'En Proceso']
                ).exists()
                if tiene_pedidos_vendedor:
                    messages.error(request, f"No puedes cambiar el rol de {usuario_editar.nombres}. Tiene locales con pedidos pendientes por entregar.")
                    return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

            # 2. Si es Cliente y tiene pedidos realizados
            if usuario_editar.rol.nombre_rol == "Cliente":
                tiene_pedidos_cliente = Pedido.objects.filter(
                    usuario=usuario_editar,
                    estado__estado_pedido__in=['PENDIENTE', 'En Proceso']
                ).exists()
                if tiene_pedidos_cliente:
                    messages.error(request, f"{usuario_editar.nombres} tiene pedidos en proceso como cliente. Debe finalizarlos antes de cambiar su rol.")
                    return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

            usuario_editar.rol = nuevo_rol
            usuario_editar.save()
            
            # Si el admin se cambia el rol a sí mismo, actualizamos su sesión o lo sacamos
            if str(usuario_editar.id) == str(request.session.get("usuario_id")):
                request.session["usuario_rol"] = nuevo_rol.nombre_rol
                if nuevo_rol.nombre_rol != "Administrador":
                    messages.info(request, "Tu rol ha cambiado. Acceso administrativo revocado.")
                    return redirect("landing")

            messages.success(request, f"El rol de {usuario_editar.nombres} ahora es {nuevo_rol.nombre_rol}.")
            return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

        elif accion == "eliminar_local":
            local_id = request.POST.get("local_id")
            local = get_object_or_404(Local, IdLocal=local_id)
            motivo = request.POST.get("motivo", "").strip() or "No se especificó un motivo."
            
            with transaction.atomic():
                if Pedido.objects.filter(detallepedido__prenda__idLocal=local, estado__estado_pedido__in=['PENDIENTE', 'En Proceso']).exists():
                    messages.error(request, "No se puede eliminar: el local tiene pedidos en curso.")
                else:
                    # Notificar al vendedor antes de borrar el objeto local
                    Notificacion.objects.create(
                        usuario=local.IdUsuario,
                        mensaje=f"Tu local '{local.Nombre_local}' ha sido eliminado por la administración.",
                        motivo=motivo,
                        tipo='error'
                    )

                    # Guardar datos para deshacer
                    request.session['last_deleted_local'] = {
                        'IdUsuario_id': local.IdUsuario_id,
                        'Nombre_local': local.Nombre_local,
                        'Descripcion': local.Descripcion,
                        'Ubicacion_direccion': local.Ubicacion_direccion,
                        'Imagen': local.Imagen.name if local.Imagen else None,
                        'Horaapertura': str(local.Horaapertura) if local.Horaapertura else None,
                        'HoraCierre': str(local.HoraCierre) if local.HoraCierre else None,
                        'EstaActivo': local.EstaActivo,
                        'Numero': local.Numero
                    }
                    nombre = local.Nombre_local
                    local.delete()
                    messages.success(request, f"Local <b>{nombre}</b> eliminado. ¿Deseas deshacer? <form method='POST' class='inline ml-2'><input type='hidden' name='csrfmiddlewaretoken' value='{get_token(request)}'><input type='hidden' name='accion' value='deshacer'><button type='submit' class='underline font-black uppercase text-emerald-700 hover:text-emerald-900'>[Deshacer]</button></form> <span class='countdown-timer ml-2 opacity-50'>(10s)</span>")
            return redirect(f"{reverse('admin_dashboard')}?section=locales")

        elif accion == "eliminar_prenda":
            prenda_id = request.POST.get("prenda_id")
            prenda = get_object_or_404(Prendas, pk=prenda_id)
            motivo = request.POST.get("motivo", "").strip() or "Incumplimiento de las políticas de la plataforma."
            
            with transaction.atomic():
                # Validación de integridad: No borrar si hay pedidos pendientes que dependan de este ID exacto
                if Pedido.objects.filter(detallepedido__prenda=prenda, estado__estado_pedido__in=['PENDIENTE', 'En Proceso']).exists():
                    messages.error(request, f"No se puede eliminar permanentemente: el uniforme '{prenda.nombre}' tiene pedidos en curso.")
                else:
                    # Notificar al vendedor (Dueño del local)
                    Notificacion.objects.create(
                        usuario=prenda.idLocal.IdUsuario,
                        mensaje=f"Tu uniforme '{prenda.nombre}' ha sido eliminado permanentemente por la administración.",
                        motivo=motivo,
                        tipo='error'
                    )

                    # Respaldar en sesión para deshacer
                    request.session['last_deleted_prenda'] = {
                        'idLocal_id': prenda.idLocal_id,
                        'nombre': prenda.nombre,
                        'descripcion': prenda.descripcion,
                        'precio': str(prenda.precio),
                        'talla': prenda.talla,
                        'material': prenda.material,
                        'stock': prenda.stock,
                        'activo': prenda.activo,
                        'tipoPrenda': prenda.tipoPrenda,
                        'imagen': prenda.imagen.name if prenda.imagen else None
                    }
                    nombre = prenda.nombre
                    prenda.delete()
                    messages.success(request, f"Uniforme <b>{nombre}</b> eliminado permanentemente. ¿Deseas deshacer? <form method='POST' class='inline ml-2'><input type='hidden' name='csrfmiddlewaretoken' value='{get_token(request)}'><input type='hidden' name='accion' value='deshacer'><button type='submit' class='underline font-black uppercase text-emerald-700 hover:text-emerald-900'>[Deshacer]</button></form> <span class='countdown-timer ml-2 opacity-50'>(10s)</span>")
            return redirect(f"{reverse('admin_dashboard')}?section=uniformes")

        elif accion == "marcar_revisado":
            com_id = request.POST.get("comentario_id")
            comentario = get_object_or_404(CalificacionPedido, pk=com_id)
            comentario.estado_moderacion = 'REVISADO'
            comentario.esta_visible = True
            comentario.save()
            messages.success(request, "Comentario aprobado.")
            return redirect(f"{reverse('admin_dashboard')}?section=comentarios")

        elif accion == "advertir_comentario":
            com_id = request.POST.get("comentario_id")
            comentario = get_object_or_404(CalificacionPedido, pk=com_id)
            comentario.estado_moderacion = 'ADVERTIDO'
            comentario.save()
            Notificacion.objects.create(
                usuario=comentario.usuario,
                mensaje="Hemos detectado actividades o comentarios que podrían infringir las políticas de convivencia y respeto de UniSena. Te recomendamos mantener un lenguaje adecuado y respetuoso dentro de la plataforma. La reincidencia en este tipo de comportamientos puede ocasionar restricciones temporales o permanentes en tu cuenta.",
                tipo='alerta'
            )
            messages.success(request, "Advertencia enviada.")
            return redirect(f"{reverse('admin_dashboard')}?section=comentarios")

        elif accion == "bloquear_usuario_moderacion":
            com_id = request.POST.get("comentario_id")
            plazo = request.POST.get("plazo_bloqueo")
            comentario = get_object_or_404(CalificacionPedido, pk=com_id)
            usuario = comentario.usuario
            
            if plazo == "perm":
                usuario.delete()
                messages.success(request, "Cuenta eliminada permanentemente.")
            else:
                # Calcular tiempo
                tiempos = {'30s': 30, '1d': 86400, '3d': 259200, '1w': 604800, '1m': 2592000}
                labels_tiempos = {'30s': '30 segundos', '1d': '1 día', '3d': '3 días', '1w': '1 semana', '1m': '1 mes'}
                segundos = tiempos.get(plazo, 30)
                label_legible = labels_tiempos.get(plazo, plazo)
                
                # Limpiar la razón de sanciones previas para evitar duplicados (ej: SANCION 1 DIA: SANCION 30s...)
                razon_limpia = comentario.motivo_deteccion or "Lenguaje inapropiado"
                if ":" in razon_limpia:
                    razon_limpia = razon_limpia.split(":")[-1].strip()

                usuario.bloqueado_hasta = now() + timedelta(seconds=segundos)
                usuario.motivo_bloqueo = razon_limpia
                usuario.save()
                # Enviar notificación al usuario bloqueado
                Notificacion.objects.create(
                    usuario=usuario,
                    mensaje=f"Tu cuenta ha sido suspendida por {label_legible}. Motivo: {razon_limpia}. La reincidencia podria causar la eliminacion permanente.",
                    tipo='error'
                )

                comentario.estado_moderacion = 'BLOQUEADO'
                comentario.motivo_deteccion = f"SANCION {label_legible.upper()}: {razon_limpia}"
                comentario.esta_visible = False
                comentario.save()
                messages.success(request, f"Usuario bloqueado por {plazo}.")
            return redirect(request.META.get('HTTP_REFERER', f"{reverse('admin_dashboard')}?section=comentarios"))

        elif accion == "bloquear_usuario":
            target_user_id = request.POST.get("user_id")
            plazo = request.POST.get("plazo_bloqueo")
            motivo = request.POST.get("motivo", "").strip() or "Incumplimiento de las normas de la plataforma."
            
            usuario = get_object_or_404(Usuario, id=target_user_id)
            
            if plazo == "perm":
                usuario.delete()
                messages.success(request, "Cuenta eliminada permanentemente.")
            else:
                # Calcular tiempo
                tiempos = {'30s': 30, '1d': 86400, '3d': 259200, '1w': 604800, '1m': 2592000}
                labels_tiempos = {'30s': '30 segundos', '1d': '1 día', '3d': '3 días', '1w': '1 semana', '1m': '1 mes'}
                segundos = tiempos.get(plazo, 30)
                label_legible = labels_tiempos.get(plazo, plazo)
                
                usuario.bloqueado_hasta = now() + timedelta(seconds=segundos)
                usuario.motivo_bloqueo = motivo
                usuario.save()
                
                # Enviar notificación al usuario bloqueado
                Notificacion.objects.create(
                    usuario=usuario,
                    mensaje=f"Tu cuenta ha sido suspendida por {label_legible}. Motivo: {motivo}. La reincidencia podria causar la eliminacion permanente.",
                    tipo='error'
                )
                messages.success(request, f"Usuario bloqueado por {plazo}.")
            return redirect(request.META.get('HTTP_REFERER', f"{reverse('admin_dashboard')}?section=usuarios"))

        elif accion == "enviar_notificacion":
            user_id = request.POST.get("user_id")
            mensaje = request.POST.get("mensaje", "").strip()
            tipo_notif = request.POST.get("tipo_notif", "info")

            if not user_id or not mensaje:
                messages.error(request, "Se requiere un destinatario y un mensaje.")
            elif len(mensaje) > 500:
                messages.error(request, "El mensaje no puede exceder los 500 caracteres.")
            else:
                usuario_notificar = get_object_or_404(Usuario, id=user_id)
                Notificacion.objects.create(
                    usuario=usuario_notificar,
                    mensaje=mensaje,
                    tipo=tipo_notif
                )
                messages.success(request, f"Notificación enviada correctamente a {usuario_notificar.nombres}.")
            return redirect(f"{reverse('admin_dashboard')}?section=usuarios")

    # --- LÓGICA DE DATOS Y FILTROS ---
    q_search = request.GET.get('q_search', '').strip()
    
    # Notificaciones para el Admin
    # Nota: Se filtran pero ya no se mostrarán en el dashboard según requerimiento
    notifs = Notificacion.objects.filter(usuario_id=request.session.get("usuario_id")).order_by('-fecha')[:10]
    
    # --- Lógica para el Dashboard Analítico ---
    dashboard_data = {}
    if section == 'dashboard':
        # 1. Ventas Totales (solo pedidos completados)
        # Usaremos esta query base para todos los cálculos de ventas
        completed_details_qs = DetallePedido.objects.filter(pedido__estado__estado_pedido='COMPLETADO')
        total_sales = completed_details_qs.aggregate(total=Sum('total_pedido'))['total'] or 0
        dashboard_data['total_sales'] = total_sales

        # 2. Conteo de Usuarios por Rol
        dashboard_data['total_users'] = Usuario.objects.count()
        dashboard_data['total_clients'] = Usuario.objects.filter(rol__nombre_rol='Cliente').count()
        dashboard_data['total_vendors'] = Usuario.objects.filter(rol__nombre_rol='Vendedor').count()
        dashboard_data['total_admins'] = Usuario.objects.filter(rol__nombre_rol='Administrador').count()
        
        # 3. Tarjetas de "Top"
        top_vendor_obj = Usuario.objects.filter(rol__nombre_rol='Vendedor', locales__prendas__detallepedido__pedido__estado__estado_pedido='COMPLETADO').annotate(
            total_vendido=Sum('locales__prendas__detallepedido__total_pedido')
        ).order_by('-total_vendido').first()
        dashboard_data['top_vendor'] = {
            'nombres': f"{top_vendor_obj.nombres} {top_vendor_obj.apellidos}" if top_vendor_obj else None,
            'total_vendido': top_vendor_obj.total_vendido if top_vendor_obj else 0
        }

        top_local_obj = Local.objects.filter(prendas__detallepedido__pedido__estado__estado_pedido='COMPLETADO').annotate(
            total_vendido=Sum('prendas__detallepedido__total_pedido')
        ).order_by('-total_vendido').first()
        dashboard_data['top_local'] = {
            'Nombre_local': top_local_obj.Nombre_local if top_local_obj else None,
            'total_vendido': top_local_obj.total_vendido if top_local_obj else 0
        }

        top_client_obj = Usuario.objects.filter(rol__nombre_rol='Cliente', pedidos__estado__estado_pedido='COMPLETADO').annotate(
            total_comprado=Sum('pedidos__detallepedido__total_pedido')
        ).order_by('-total_comprado').first()
        dashboard_data['top_client'] = {
            'nombres': f"{top_client_obj.nombres} {top_client_obj.apellidos}" if top_client_obj else None,
            'total_comprado': top_client_obj.total_comprado if top_client_obj else 0
        }

        # 4. Datos para Gráficos
        # Gráfico 1: Ventas por Mes (últimos 12 meses)
        sales_period = request.GET.get('sales_period', 'month')
        
        if sales_period == 'day':
            trunc_func, start_date, date_format = TruncDay, now() - timedelta(days=30), "%d %b"
        elif sales_period == 'year':
            trunc_func, start_date, date_format = TruncYear, now() - timedelta(days=5*365), "%Y"
        else: # 'month' por defecto
            trunc_func, start_date, date_format = TruncMonth, now() - timedelta(days=365), "%b %Y"

        # --- SOLUCIÓN DEFINITIVA: Usar tzinfo para evitar CONVERT_TZ en MySQL ---
        # Obtenemos la zona horaria actual de Django para pasársela a la función Trunc.
        # Esto hace que la truncación sea consciente de la zona horaria a nivel de aplicación,
        # no de base de datos, resolviendo el problema de `CONVERT_TZ` devolviendo NULL.
        # --- CORRECCIÓN: Consulta optimizada desde DetallePedido ---
        # 1. Filtramos los detalles de pedidos completados en el rango de fechas.
        # 2. Anotamos el período truncado (día, mes o año) desde la fecha del pedido.
        # 3. Usamos .values() para agrupar por ese período.
        # 4. Anotamos la suma de los 'total_pedido' de los detalles.
        sales_by_period = DetallePedido.objects.filter(
            pedido__estado__estado_pedido='COMPLETADO',
            pedido__fecha_pedido__gte=start_date
        ).annotate(
            period=trunc_func('pedido__fecha_pedido', output_field=DateField(), tzinfo=now().tzinfo)
        ).values('period').annotate(
            total=Sum('total_pedido')
        ).order_by('period')
        
        print(f"--- DEBUG: Período de ventas: {sales_period} ---")
        print(f"Consulta SQL generada: {sales_by_period.query}")

        sales_chart_data = []
        for s in sales_by_period:
            if s["period"] is not None and s["total"] is not None:
                sales_chart_data.append({
                    "label": s["period"].strftime(date_format),
                    "total": float(s["total"] or 0)
                })
        
        # --- DEBUGGING ADICIONAL ---
        print("========== DEBUG DASHBOARD ==========")
        print("Pedidos completados (conteo):", Pedido.objects.filter(estado__estado_pedido='COMPLETADO').count())
        print("Detalles de pedidos completados (conteo):", completed_details_qs.count())
        print("Total ventas (agregado):", completed_details_qs.aggregate(Sum('total_pedido')))
        print("--- Primeros 10 detalles de pedidos completados ---")
        for d in completed_details_qs.select_related('pedido', 'pedido__estado')[:10]:
            print(
                f"ID Detalle: {d.idDetalle}, "
                f"ID Pedido: {d.pedido.idPedido}, "
                f"Estado: {d.pedido.estado.estado_pedido}, "
                f"Fecha: {d.pedido.fecha_pedido}, "
                f"Total: {d.total_pedido}"
            )
        print("======================================")

        print(f"Datos procesados para el gráfico: {sales_chart_data}")

        dashboard_data['sales_chart'] = {
            "labels": [x["label"] for x in sales_chart_data],
            "data": [x["total"] for x in sales_chart_data]
        }
        # Pasamos el período seleccionado al contexto para el botón activo
        dashboard_data['sales_period'] = sales_period

        # Gráfico 2: Pedidos por Estado
        orders_by_status = Pedido.objects.values('estado__estado_pedido').annotate(count=Count('idPedido')).order_by('-count')
        dashboard_data['orders_chart'] = {
            'labels': [s['estado__estado_pedido'] for s in orders_by_status],
            'data': [s['count'] for s in orders_by_status]
        }

        # Gráfico 3: Usuarios por Rol
        dashboard_data['users_chart'] = {
            'labels': ['Clientes', 'Vendedores', 'Admins'],
            'data': [dashboard_data['total_clients'], dashboard_data['total_vendors'], dashboard_data['total_admins']]
        }

        # Gráfico 4: Top 5 Uniformes más vendidos (por cantidad)
        top_uniforms = Prendas.objects.annotate(
            units_sold=Sum('detallepedido__cantidad', filter=Q(detallepedido__pedido__estado__estado_pedido='COMPLETADO'))
        ).filter(units_sold__gt=0).order_by('-units_sold')[:5]
        dashboard_data['top_uniforms_chart'] = {
            'labels': [u.nombre for u in top_uniforms],
            'data': [u.units_sold for u in top_uniforms]
        }

        # Serializar datos del dashboard para JS
        dashboard_data_json = json.dumps(dashboard_data, cls=DjangoJSONEncoder)

    usuarios, locales, pedidos, uniformes, comentarios = [], [], [], [], []
    
    if section == 'usuarios':
        usuarios = Usuario.objects.all().select_related('rol').order_by('-id')
        f_rol = request.GET.get('f_rol', '')
        f_tipo_id = request.GET.get('f_tipo_id', '')
        f_status_user = request.GET.get('f_status_user', '')

        if q_search:
            usuarios = usuarios.filter(Q(nombres__icontains=q_search) | Q(apellidos__icontains=q_search) | Q(num_identificacion__icontains=q_search) | Q(correo__icontains=q_search))
        if f_rol: usuarios = usuarios.filter(rol_id=f_rol)
        if f_tipo_id: usuarios = usuarios.filter(tipo_identificacion=f_tipo_id)
        if f_status_user == 'blocked':
            # Usuarios con bloqueo activo o pendientes de aceptar políticas
            usuarios = usuarios.filter(Q(bloqueado_hasta__gt=now()) | Q(pendiente_aceptar_politicas=True))
        elif f_status_user == 'active':
            usuarios = usuarios.filter(Q(bloqueado_hasta__isnull=True) | Q(bloqueado_hasta__lte=now()), pendiente_aceptar_politicas=False)

        # Añadir datos para el modal de gestión de usuario
        usuarios = usuarios.annotate(
            is_blocked=Q(bloqueado_hasta__gt=now()) | Q(pendiente_aceptar_politicas=True)
        )
        # Convertir a lista de diccionarios para serializar a JSON
        usuarios_json = list(usuarios.values('id', 'nombres', 'apellidos', 'rol__nombre_rol', 'is_blocked', 'bloqueado_hasta', 'motivo_bloqueo', 'rol_id'))
        usuarios_json_str = json.dumps(usuarios_json, cls=DjangoJSONEncoder)

    elif section == 'locales':
        locales = Local.objects.all().select_related('IdUsuario').order_by('-IdLocal')
        f_status = request.GET.get('f_status', '')
        f_user_status = request.GET.get('f_user_status', '')
        if q_search:
            locales = locales.filter(Q(Nombre_local__icontains=q_search) | Q(IdUsuario__nombres__icontains=q_search) | Q(Descripcion__icontains=q_search))
        if f_status == '1': locales = locales.filter(EstaActivo=True)
        elif f_status == '0': locales = locales.filter(EstaActivo=False)
        if f_user_status == 'blocked':
            locales = locales.filter(IdUsuario__bloqueado_hasta__gt=now())
        elif f_user_status == 'active':
            locales = locales.filter(Q(IdUsuario__bloqueado_hasta__isnull=True) | Q(IdUsuario__bloqueado_hasta__lte=now()))

    elif section == 'pedidos':
        # Usamos Subquery y Sum para obtener la información detallada que solicitaste
        # Subconsulta para obtener el motivo de la notificación de cancelación si existe
        motivo_subquery = Notificacion.objects.filter(
            mensaje__icontains=OuterRef('idPedido'),
            tipo='error'
        ).values('motivo')[:1]
        detalles_qs = DetallePedido.objects.filter(pedido=OuterRef('pk'))
        primer_local = detalles_qs.values('prenda__idLocal__Nombre_local')[:1]
        primer_uniforme = detalles_qs.values('prenda__nombre')[:1]
        
        pedidos = Pedido.objects.all().select_related('usuario', 'estado').annotate(
            # Optimizamos para obtener todos los detalles de una vez
            # local_nombre=Subquery(primer_local), # Ya no es necesario, lo obtenemos del prefetch
            uniforme_nombre=Subquery(primer_uniforme),
            motivo_cancelacion=Subquery(motivo_subquery),
            cantidad_total=Sum('detallepedido__cantidad'),
            precio_total=Sum('detallepedido__total_pedido')
        ).order_by('-idPedido')
        
        f_estado = request.GET.get('f_estado', '')
        if q_search:
            pedidos = pedidos.filter(Q(idPedido__icontains=q_search) | Q(usuario__nombres__icontains=q_search) | Q(detallepedido__prenda__idLocal__Nombre_local__icontains=q_search)).distinct()
        if f_estado:
            pedidos = pedidos.filter(estado__estado_pedido=f_estado)

        # Prefetch para el nuevo modal de detalles
        pedidos = pedidos.prefetch_related('detallepedido_set__prenda__idLocal__IdUsuario')


    elif section == 'uniformes':
        uniformes = Prendas.objects.all().select_related('idLocal', 'idLocal__IdUsuario').order_by('-idPrenda')
        f_talla = request.GET.get('f_talla', '')
        f_material = request.GET.get('f_material', '')
        f_tipo_prenda = request.GET.get('f_tipo_prenda', '')
        if q_search:
            uniformes = uniformes.filter(Q(nombre__icontains=q_search) | Q(descripcion__icontains=q_search) | Q(idPrenda__icontains=q_search) | Q(idLocal__Nombre_local__icontains=q_search))
        if f_talla: uniformes = uniformes.filter(talla=f_talla)
        if f_material: uniformes = uniformes.filter(material=f_material)
        if f_tipo_prenda: uniformes = uniformes.filter(tipoPrenda=f_tipo_prenda)

    elif section == 'comentarios':
        # Mostramos todos los comentarios sin filtros (cola completa para revisión manual)
        comentarios = CalificacionPedido.objects.all().select_related('usuario', 'locales').order_by('-fecha')
        if q_search:
            comentarios = comentarios.filter(
                Q(comentario__icontains=q_search) | 
                Q(usuario__nombres__icontains=q_search) | 
                Q(locales__Nombre_local__icontains=q_search) |
                Q(motivo_deteccion__icontains=q_search)
            )
        f_estado_moderacion = request.GET.get('f_estado_moderacion', '')
        if f_estado_moderacion:
            comentarios = comentarios.filter(estado_moderacion=f_estado_moderacion
            )

    context = {
        "nombre": request.session.get("usuario_nombre"),
        "section": section,
        "usuarios": usuarios,
        "locales": locales,
        "dashboard_data_json": dashboard_data_json if section == 'dashboard' else '{}',
        "dashboard_data": dashboard_data,
        "pedidos": pedidos,
        "uniformes": uniformes,
        "comentarios": comentarios,
        "roles": Rol.objects.all(),
        "notificaciones": notifs,
        "estados_pedido": EstadoPedido.ESTADO_CHOICES,
        "tipo_identificacion_choices": Usuario.TIPOS_IDENTIFICACION,
        "q_search": q_search,
        "f_rol_sel": request.GET.get('f_rol', ''),
        "f_tipo_id_sel": request.GET.get('f_tipo_id', ''),
        "f_status_user_sel": request.GET.get('f_status_user', ''),
        "f_status_sel": request.GET.get('f_status', ''),
        "f_user_status_sel": request.GET.get('f_user_status', ''),
        "f_estado_sel": request.GET.get('f_estado', ''),
        "f_talla_sel": request.GET.get('f_talla', ''),
        "f_material_sel": request.GET.get('f_material', ''),
        "f_tipo_prenda_sel": request.GET.get('f_tipo_prenda', ''),
        "f_estado_moderacion_sel": request.GET.get('f_estado_moderacion', ''),
        "now": now(), # Pasar la hora actual al template para comparaciones
        "timezone": "America/Bogota", # Para consistencia en el template si es necesario
        "talla_choices": Prendas.TALLA_CHOICES,
        "material_choices": Prendas.MATERIAL_CHOICES,
        "prenda_tipo_choices": Prendas.TIPO_PRENDA_CHOICES,
        "usuarios_json": usuarios_json_str if section == 'usuarios' else '[]',
    }

    if section == 'reportes':
        report_type = request.GET.get('report_type', 'ventas')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        locales_filter = request.GET.getlist('locales')
        estado_filter = request.GET.get('estado', '')
        export_format = request.GET.get('export')
        
        # Convertir a enteros para la comparación en el template
        locales_filter_int = [int(l) for l in locales_filter]

        # 1. Validación de Fechas
        if fecha_inicio and fecha_fin and fecha_inicio > fecha_fin:
            messages.error(request, "La fecha de inicio no puede ser mayor que la fecha final.")
            context.update({
                'report_type': report_type,
                'headers': [],
                'data': [],
                'total_records': 0,
                'f_fecha_inicio': fecha_inicio,
                'f_fecha_fin': fecha_fin,
                'f_locales_sel': locales_filter_int,
                'f_estado_sel': estado_filter,
            })
            return render(request, "dashadmin.html", context)

        data = []
        headers = []
        
        if report_type == 'ventas':
            qs = DetallePedido.objects.select_related('pedido__usuario', 'prenda__idLocal')
            print("--- DEBUG: Reporte de Ventas ---")
            print("request.GET reportes:", dict(request.GET))
            print("Antes:", qs.count())

            # --- Lógica de Fechas con Zona Horaria ---
            if fecha_inicio:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                inicio = now().replace(year=fecha_inicio_obj.year, month=fecha_inicio_obj.month, day=fecha_inicio_obj.day, hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(pedido__fecha_pedido__gte=inicio)
                print("Fecha inicio (aware):", inicio)

            if fecha_fin:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                fin = now().replace(year=fecha_fin_obj.year, month=fecha_fin_obj.month, day=fecha_fin_obj.day, hour=23, minute=59, second=59, microsecond=999999)
                qs = qs.filter(pedido__fecha_pedido__lte=fin)
                print("Fecha fin (aware):", fin)
            
            print("Después de filtros de fecha:", qs.count())

            headers = ['Fecha', 'Pedido ID', 'Cliente', 'Local', 'Total']
            if estado_filter:
                qs = qs.filter(pedido__estado__estado_pedido=estado_filter)
            else:
                qs = qs.filter(pedido__estado__estado_pedido='COMPLETADO')
            print("Después estado:", qs.count())
            if locales_filter:
                qs = qs.filter(prenda__idLocal_id__in=locales_filter)
            print("Después local:", qs.count())

            qs = qs.order_by('-pedido__fecha_pedido')
            print("SQL:", str(qs.query))

            data = list(qs.values('pedido__fecha_pedido', 'pedido__idPedido', 'pedido__usuario__nombres', 'prenda__idLocal__Nombre_local', 'total_pedido'))

        elif report_type == 'pedidos':
            qs = Pedido.objects.select_related('usuario', 'estado').annotate(total_general=Sum('detallepedido__total_pedido'))
            print("--- DEBUG: Reporte de Pedidos ---")
            print("request.GET reportes:", dict(request.GET))
            print("Antes:", qs.count())

            if fecha_inicio:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                inicio = now().replace(year=fecha_inicio_obj.year, month=fecha_inicio_obj.month, day=fecha_inicio_obj.day, hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(fecha_pedido__gte=inicio)

            if fecha_fin:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                fin = now().replace(year=fecha_fin_obj.year, month=fecha_fin_obj.month, day=fecha_fin_obj.day, hour=23, minute=59, second=59, microsecond=999999)
                qs = qs.filter(fecha_pedido__lte=fin)
            print("Después de filtros de fecha:", qs.count())

            headers = ['Pedido ID', 'Cliente', 'Estado', 'Fecha', 'Total']

            if estado_filter:
                qs = qs.filter(estado__estado_pedido=estado_filter)
            print("Después estado:", qs.count())

            if locales_filter:
                qs = qs.filter(detallepedido__prenda__idLocal_id__in=locales_filter).distinct()
            print("Después local:", qs.count())

            qs = qs.order_by('-fecha_pedido')
            print("SQL:", str(qs.query))

            # Log de depuración de fechas en la BD
            print("--- DEBUG: Fechas de Pedidos en BD (UTC) ---")
            print(list(Pedido.objects.values_list("idPedido", "fecha_pedido")))
            print("-----------------------------------------")
            
            data = list(qs.values('idPedido', 'usuario__nombres', 'estado__estado_pedido', 'fecha_pedido', 'total_general'))

        elif report_type == 'locales': # REPORTE DE LOCALES
            headers = ['Local', 'Vendedor', 'Estado', 'Productos Registrados']
            qs = Local.objects.select_related('IdUsuario').annotate(
                num_productos=Count('prendas')
            ).order_by('Nombre_local')
            print(f"--- DEBUG: Reporte de Locales ---")

            # Aplicar filtros de fecha si existen (asumiendo que Local tiene un campo de fecha de creación)
            if fecha_inicio:
                # Asumimos que el modelo Local tiene un campo `fecha_creacion` o similar.
                # Si no lo tiene, este filtro no hará nada. Lo añadimos para que sea consistente.
                # Usamos __date para comparar solo la fecha.
                qs = qs.filter(fecha_creacion__date__gte=fecha_inicio)
            if fecha_fin:
                qs = qs.filter(fecha_creacion__date__lte=fecha_fin)

            if locales_filter: qs = qs.filter(IdLocal__in=locales_filter)
            
            # Procesar datos para mostrar 'Activo'/'Inactivo'
            data = []
            for local in qs:
                data.append({'Nombre_local': local.Nombre_local, 'Vendedor': local.IdUsuario.nombres, 'Estado': 'Activo' if local.EstaActivo else 'Inactivo', 'Productos': local.num_productos})

        elif report_type == 'uniformes':
            headers = ['Prenda', 'Local', 'Stock', 'Precio']
            qs = Prendas.objects.select_related('idLocal').order_by('nombre')
            print(f"--- DEBUG: Reporte de Uniformes ---")
            if locales_filter: qs = qs.filter(idLocal_id__in=locales_filter)
            data = list(qs.values('nombre', 'idLocal__Nombre_local', 'stock', 'precio'))

        elif report_type == 'usuarios':
            f_status_user = request.GET.get('f_status_user', '') # Nuevo filtro de estado
            headers = ['Nombre', 'Rol', 'Correo', 'Estado']
            qs = Usuario.objects.select_related('rol').order_by('nombres')
            
            if f_status_user == 'blocked':
                qs = qs.filter(bloqueado_hasta__gt=now())
            elif f_status_user == 'restricted':
                qs = qs.filter(pendiente_aceptar_politicas=True)
            elif f_status_user == 'active':
                qs = qs.filter(Q(bloqueado_hasta__isnull=True) | Q(bloqueado_hasta__lte=now()), pendiente_aceptar_politicas=False)

            data = []
            print(f"--- DEBUG: Reporte de Usuarios ---")
            for u in qs:
                if u.bloqueado_hasta and u.bloqueado_hasta > now(): estado = "Bloqueado"
                elif u.pendiente_aceptar_politicas: estado = "Restringido"
                else: estado = "Activo"
                data.append({'nombres': f"{u.nombres} {u.apellidos}", 'rol__nombre_rol': u.rol.nombre_rol, 'correo': u.correo, 'estado': estado})

        if export_format == 'excel':
            return generar_excel(report_type, headers, data)
        if export_format == 'pdf':
            return generar_pdf(report_type, headers, data)

        context.update({
            'report_type': report_type,
            'headers': headers,
            'data': data[:100],
            'total_records': len(data),
            'f_fecha_inicio': fecha_inicio,
            'f_fecha_fin': fecha_fin,
            'f_locales_sel': locales_filter_int,
            'f_estado_sel': estado_filter,
            'f_status_user_sel': f_status_user, # Pasar el nuevo filtro al contexto
            'all_locales_for_filter': Local.objects.all().order_by('Nombre_local'), # 2. Cargar locales para el filtro
        })

    return render(request, "dashadmin.html", context)

@never_cache
def politicas(request):
    return render(request, "politicas.html")

def reportes_view(request):
    """Vista dedicada para el Centro de Reportes."""

def generar_excel(report_name, headers, data):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_{report_name}_{now().strftime("%Y-%m-%d")}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_name.capitalize()

    # Estilo para cabeceras
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="125F58", end_color="125F58", fill_type="solid")

    # Escribir cabeceras
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Escribir datos
    for row_num, row_data in enumerate(data, 2):
        # Usamos los headers para obtener los valores en el orden correcto
        values_list = list(row_data.values())
        for col_num, value in enumerate(values_list, 1):
            # Formatear fechas y decimales
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")
            elif isinstance(value, datetime.date): # Asumiendo que 'date' se importa de 'datetime'
                value = value.strftime("%Y-%m-%d")
            elif isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-ajustar columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

def generar_pdf(report_name, headers, data):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Título y cabecera
    p.setFont("Helvetica-Bold", 16)
    p.setFillColorRGB(0.07, 0.37, 0.35) # #125f58
    p.drawString(inch, height - inch, f"UniSena - Reporte de {report_name.capitalize()}")
    
    p.setFont("Helvetica", 9)
    p.setFillColorRGB(0.4, 0.4, 0.4)
    p.drawString(inch, height - inch - 20, f"Generado el: {now().strftime('%d/%m/%Y %H:%M')} | Total de registros: {len(data)}")
    
    # Línea separadora
    p.line(inch, height - inch - 30, width - inch, height - inch - 30)

    # Tabla
    p.setFont("Helvetica-Bold", 10)
    p.setFillColorRGB(0, 0, 0)
    y = height - inch - 60
    x_positions = [inch]
    col_widths = [(width - 2 * inch) / len(headers)] * len(headers)
    for i in range(len(headers)):
        x_positions.append(x_positions[-1] + col_widths[i])

    for i, header in enumerate(headers):
        p.drawString(x_positions[i] + 5, y, header)

    p.setFont("Helvetica", 8)
    y -= 20

    for row_data in data:
        if y < inch: # Salto de página si no hay espacio
            p.showPage()
            p.setFont("Helvetica", 8)
            y = height - inch

        values = list(row_data.values())
        for i, value in enumerate(values):
            # Formatear valores para PDF
            if isinstance(value, datetime):
                text = value.strftime("%d/%m/%y %H:%M")
            elif isinstance(value, datetime.date): # Asumiendo que 'date' se importa de 'datetime'
                text = value.strftime("%d/%m/%Y")
            elif isinstance(value, Decimal):
                text = f"${value:,.0f}"
            else:
                text = str(value)
            
            p.drawString(x_positions[i] + 5, y, text[:40]) # Limitar longitud de texto por celda
        y -= 15

    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')
